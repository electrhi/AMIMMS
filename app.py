import os
import io
import base64
import json
import ssl
import certifi
import requests
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, session, url_for, send_file
from PIL import Image, ImageDraw, ImageFont
from google.oauth2.service_account import Credentials
from google.cloud import storage
import gspread
from openpyxl import Workbook

# =========================================================
# ✅ SSL 안정화 (Render 환경용)
# =========================================================
ssl._create_default_https_context = ssl._create_unverified_context
requests.adapters.DEFAULT_RETRIES = 5

# =========================================================
# ✅ Flask 초기화
# =========================================================
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "kdn_secret_key")

# =========================================================
# ✅ Google Sheets & GCS 연결 설정
# =========================================================
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
CREDS = Credentials.from_service_account_info(
    json.loads(os.getenv("GOOGLE_CREDENTIALS_JSON")), scopes=SCOPES
)
gc = gspread.authorize(CREDS)
gc.session.verify = certifi.where()

USERS_SHEET_KEY = os.getenv("GOOGLE_USERS_SHEET_KEY")
RECORDS_SHEET_KEY = os.getenv("GOOGLE_RECORDS_SHEET_KEY")
GCS_BUCKET_NAME = os.getenv("GCS_BUCKET_NAME", "amimms-receipts")

users_sheet = gc.open_by_key(USERS_SHEET_KEY).sheet1
records_sheet = gc.open_by_key(RECORDS_SHEET_KEY).sheet1

# =========================================================
# ✅ 로그인
# =========================================================
@app.route("/", methods=["GET", "POST"])
def login():
    """사용자 로그인 (시트 기반 ID/PW 검증)"""
    df = pd.DataFrame(users_sheet.get_all_records())
    users = {
        row["ID"]: {
            "PASSWORD": row["PASSWORD"],
            "AUTHORITY": row.get("AUTHORITY", 0)
        } for _, row in df.iterrows()
    }

    if request.method == "POST":
        user_id = request.form.get("user_id", "").strip()
        pw = request.form.get("password", "").strip()
        if user_id in users and users[user_id]["PASSWORD"] == pw:
            session["logged_in"] = True
            session["user_id"] = user_id
            session["authority"] = users[user_id]["AUTHORITY"]
            return redirect(url_for("menu"))
        return render_template("login.html", error="❌ 로그인 정보가 올바르지 않습니다.")
    return render_template("login.html")

# =========================================================
# ✅ 메뉴 (로그인 사용자 표시 포함)
# =========================================================
@app.route("/menu")
def menu():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    return render_template(
        "menu.html",
        user_id=session.get("user_id", ""),
        authority=session.get("authority", 0)
    )

# =========================================================
# ✅ 자재 입력
# =========================================================
@app.route("/form", methods=["GET", "POST"])
def form():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    if request.args.get("new") == "1":
        session.pop("materials", None)

    if request.method == "POST":
        materials = []
        for i in range(len(request.form.getlist("통신방식"))):
            materials.append({
                "통신방식": request.form.getlist("통신방식")[i],
                "구분": request.form.getlist("구분")[i],
                "신철": request.form.getlist("신철")[i],
                "수량": request.form.getlist("수량")[i],
                "박스번호": request.form.getlist("박스번호")[i],
            })
        session["materials"] = materials
        return redirect(url_for("confirm"))

    return render_template("form.html", materials=session.get("materials", []))

# =========================================================
# ✅ 확인 페이지
# =========================================================
@app.route("/confirm", methods=["GET", "POST"])
def confirm():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    materials = session.get("materials", [])
    logged_user = session.get("user_id")

    if request.method == "POST":
        giver = request.form["giver"]
        receiver = request.form["receiver"]
        giver_sign = request.form["giver_sign"]
        receiver_sign = request.form["receiver_sign"]

        receipt_link = generate_receipt(materials, giver, receiver, giver_sign, receiver_sign)
        save_to_sheets(materials, giver, receiver)

        session["last_receipt"] = receipt_link
        session.pop("materials", None)
        return render_template("result.html", receipt_link=receipt_link)

    return render_template("confirm.html", materials=materials, logged_user=logged_user)
# =========================================================
# ✅ 누적 자재 현황
# =========================================================
@app.route("/summary")
def summary():
    if not session.get("logged_in"):
        return redirect(url_for("login"))

    user_id = session.get("user_id", "")
    df = pd.DataFrame(records_sheet.get_all_records())

    if df.empty or "받는사람" not in df.columns:
        return render_template("summary.html", summary_data=None, message="등록된 자재 데이터가 없습니다.")

    df = df[df["받는사람"] == user_id]
    if df.empty:
        return render_template("summary.html", summary_data=None, message="등록된 자재 데이터가 없습니다.")

    summary = df.groupby(["통신방식", "구분"], as_index=False).agg({"수량": "sum", "박스번호": "count"})
    summary.rename(columns={"수량": "합계", "박스번호": "박스수"}, inplace=True)
    summary.sort_values(["통신방식", "구분"], inplace=True)

    return render_template("summary.html", summary_data=summary.to_dict("records"))


# =========================================================
# ✅ 관리자용 종합관리표
# =========================================================
@app.route("/admin_summary")
def admin_summary():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    if session.get("authority") != 1:
        return "❌ 접근 권한이 없습니다.", 403

    user_id = session.get("user_id", "")
    df = pd.DataFrame(records_sheet.get_all_records())

    if df.empty:
        return render_template("admin_summary.html", message="등록된 자재 데이터가 없습니다.", user_id=user_id)

    df = df[df["주는사람"] == user_id]
    if df.empty:
        return render_template("admin_summary.html", message="해당 사용자의 기록이 없습니다.", user_id=user_id)

    pivot = pd.pivot_table(df, index="받는사람", columns="구분", values="수량", aggfunc="sum", fill_value=0)
    pivot.loc["합계"] = pivot.sum(numeric_only=True)
    table_html = pivot.to_html(classes="table-auto border text-center", border=1)

    return render_template("admin_summary.html", table_html=table_html, user_id=user_id)


# =========================================================
# ✅ 관리자 종합관리표 엑셀 다운로드
# =========================================================
@app.route("/download_admin_summary")
def download_admin_summary():
    if not session.get("logged_in"):
        return redirect(url_for("login"))
    if session.get("authority") != 1:
        return "❌ 접근 권한이 없습니다.", 403

    user_id = session.get("user_id", "")
    df = pd.DataFrame(records_sheet.get_all_records())

    if df.empty:
        return "❌ 다운로드할 데이터가 없습니다.", 404

    df = df[df["주는사람"] == user_id]
    if df.empty:
        return "❌ 해당 사용자의 데이터가 없습니다.", 404

    pivot = pd.pivot_table(df, index="받는사람", columns="구분", values="수량", aggfunc="sum", fill_value=0)
    pivot.loc["합계"] = pivot.sum(numeric_only=True)
    pivot.reset_index(inplace=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "종합관리표"

    for col_idx, col_name in enumerate(pivot.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(pivot.values.tolist(), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"admin_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =========================================================
# ✅ GCS 업로드 함수
# =========================================================
def upload_to_gcs(file_path, file_name, bucket_name):
    try:
        creds = Credentials.from_service_account_info(json.loads(os.environ["GOOGLE_CREDENTIALS_JSON"]))
        client = storage.Client(credentials=creds)
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(file_name)
        blob.upload_from_filename(file_path, content_type="image/jpeg")
        url = blob.generate_signed_url(expiration=3600 * 24 * 365, method="GET")
        return url
    except Exception as e:
        print(f"❌ GCS 업로드 실패: {e}")
        return None


# =========================================================
# ✅ 인수증 이미지 생성 (로고 + 한글 폰트)
# =========================================================
def generate_receipt(materials, giver, receiver, giver_sign, receiver_sign):
    width, height = 1240, 1754
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)

    # ✅ 폰트 설정
    font_path = "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"
    if not os.path.exists(font_path):
        font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"

    title_font = ImageFont.truetype(font_path, 60)
    bold_font = ImageFont.truetype(font_path, 34)

    # ✅ 로고 삽입
    logo_path = "static/kdn_logo.png"
    if os.path.exists(logo_path):
        logo = Image.open(logo_path).resize((200, 200))
        img.paste(logo, (width - 280, 80))

    draw.text((480, 100), "자재 인수증", font=title_font, fill="black")
    draw.text((100, 200), f"작성일자: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", font=bold_font, fill="black")

    # ✅ 표
    y = 300
    headers = ["통신방식", "구분", "신철", "수량", "박스번호"]
    positions = [100, 400, 600, 800, 1000]
    draw.rectangle((80, y, 1160, y + 55), outline="black", fill="#E8F0FE")

    for i, h in enumerate(headers):
        draw.text((positions[i], y + 10), h, font=bold_font, fill="black")

    y += 70
    for m in materials:
        for i, key in enumerate(headers):
            draw.text((positions[i], y), str(m.get(key, "")), font=bold_font, fill="black")
        y += 50
    draw.rectangle((80, 300, 1160, y), outline="black")

    # ✅ 서명 처리
    def decode_sign(s):
        try:
            s = s.split(",")[1] if "," in s else s
            img = Image.open(BytesIO(base64.b64decode(s)))
            return img.convert("RGBA")
        except Exception:
            return None

    giver_img, receiver_img = decode_sign(giver_sign), decode_sign(receiver_sign)
    footer_y = height - 150
    draw.text((200, footer_y - 40), f"주는 사람: {giver}", font=bold_font, fill="black")
    draw.text((800, footer_y - 40), f"받는 사람: {receiver}", font=bold_font, fill="black")

    if giver_img:
        img.paste(giver_img.resize((260, 120)), (240, footer_y - 190), giver_img)
    if receiver_img:
        img.paste(receiver_img.resize((260, 120)), (840, footer_y - 190), receiver_img)

    img.save("/tmp/receipt.jpg", "JPEG", quality=95)
    gcs_link = upload_to_gcs(
        "/tmp/receipt.jpg",
        f"receipt_{receiver}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg",
        GCS_BUCKET_NAME,
    )
    return gcs_link or "GCS 업로드 실패"


# =========================================================
# ✅ Google Sheets 저장
# =========================================================
def save_to_sheets(materials, giver, receiver):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for m in materials:
        records_sheet.append_row([
            m["통신방식"], m["구분"], giver, receiver,
            m["신철"], m["수량"], m["박스번호"], now
        ])


# =========================================================
# ✅ 인수증 다운로드
# =========================================================
@app.route("/download_receipt")
def download_receipt():
    receipt_path = session.get("last_receipt")
    if receipt_path:
        return redirect(receipt_path)
    return "❌ 인수증 파일을 찾을 수 없습니다.", 404


# =========================================================
# ✅ 로그아웃
# =========================================================
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =========================================================
# ✅ 서버 실행
# =========================================================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
